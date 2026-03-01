"""Excel export endpoints for major data tables."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import json
import re

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.api import appointments as appointments_module
from app.api import customers as customers_module
from app.api import employees as employees_module
from app.api import finance as finance_module
from app.api import settings as settings_module
from app.api import treatments as treatments_module
from app.core.middleware import require_auth

router = APIRouter(prefix="/api", tags=["exports"])


@router.get("/export/{resource}")
async def export_resource(
    resource: str,
    request: Request,
    format: str = Query(default="xlsx"),
    columns: str | None = Query(default=None),
    _user: dict = Depends(require_auth),
):
    if (format or "").strip().lower() != "xlsx":
        raise HTTPException(status_code=400, detail="Only format=xlsx is supported")

    rows = await _load_rows(resource=resource, request=request, user=_user)

    requested_columns = _parse_columns(columns)
    all_columns = requested_columns or _collect_columns(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Export"

    if not all_columns:
        all_columns = ["message"]
        rows = [{"message": "No data"}]

    header_labels = [_column_label(col) for col in all_columns]
    sheet.append(header_labels)

    header_fill = PatternFill(fill_type="solid", fgColor="E8F0FE")
    for idx in range(1, len(header_labels) + 1):
        cell = sheet.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        output_row = []
        for col in all_columns:
            value = row.get(col)
            output_row.append(_serialise_cell(value))
        sheet.append(output_row)

    for column_cells in sheet.columns:
        max_length = 10
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 52)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    safe_resource = re.sub(r"[^a-zA-Z0-9_-]+", "-", resource.strip().lower()) or "export"
    filename = f"{safe_resource}-{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _load_rows(resource: str, request: Request, user: dict) -> list[dict]:
    key = (resource or "").strip().lower()
    params = request.query_params

    if key == "customers":
        payload = customers_module.list_customers(
            page=1,
            per_page=0,
            offset=0,
            limit=0,
            search=params.get("search"),
            sort=params.get("sort"),
            order=params.get("order") or "desc",
            company=params.get("company"),
            company_id=params.get("companyId"),
            _user=user,
        )
        return _items(payload)

    if key == "appointments":
        payload = await appointments_module.list_appointments(
            page=1,
            perPage=0,
            offset=0,
            limit=0,
            dateFrom=_parse_date(params.get("dateFrom")),
            dateTo=_parse_date(params.get("dateTo")),
            companyId=params.get("companyId"),
            company=params.get("company"),
            state=params.get("state"),
            q=params.get("q") or params.get("search"),
        )
        return _items(payload)

    if key in {"sale-orders", "sale_orders", "treatments"}:
        payload = await treatments_module.list_sale_orders(
            page=1,
            per_page=0,
            offset=0,
            limit=0,
            search=params.get("search") or "",
            companyId=params.get("companyId"),
            partnerId=params.get("partnerId"),
            state=params.get("state"),
            dateFrom=_parse_date(params.get("dateFrom")),
            dateTo=_parse_date(params.get("dateTo")),
            _user=user,
        )
        return _items(payload)

    if key == "payments":
        payload = await finance_module.list_payments(
            page=1,
            per_page=0,
            offset=0,
            limit=0,
            companyId=params.get("companyId"),
            partnerId=params.get("partnerId"),
            paymentType=params.get("paymentType"),
            dateFrom=_parse_date(params.get("dateFrom")),
            dateTo=_parse_date(params.get("dateTo")),
            state=params.get("state"),
            _user=user,
        )
        return _items(payload)

    if key == "employees":
        payload = await employees_module.list_employees(
            page=1,
            per_page=0,
            offset=0,
            limit=0,
            search=params.get("search") or "",
            companyId=params.get("companyId"),
            isDoctor=_parse_bool(params.get("isDoctor")),
            _user=user,
        )
        return _items(payload)

    if key == "users":
        payload = await settings_module.list_users(
            page=1,
            per_page=0,
            offset=0,
            limit=0,
            search=params.get("search") or "",
            _user=user,
        )
        return _items(payload)

    raise HTTPException(status_code=404, detail=f"Unsupported export resource: {resource}")


def _items(payload) -> list[dict]:
    if payload is None:
        return []
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        items = payload.get("items")
        if isinstance(items, list):
            return [item for item in items if isinstance(item, dict)]
    return []


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _parse_bool(value: str | None) -> bool | None:
    if value is None:
        return None
    raw = value.strip().lower()
    if raw in {"1", "true", "yes", "y"}:
        return True
    if raw in {"0", "false", "no", "n"}:
        return False
    return None


def _parse_columns(raw: str | None) -> list[str]:
    if not raw:
        return []
    items = []
    seen = set()
    for part in raw.split(","):
        key = part.strip()
        if not key or key in seen:
            continue
        seen.add(key)
        items.append(key)
    return items


def _collect_columns(rows: list[dict]) -> list[str]:
    if not rows:
        return []
    ordered: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key in seen:
                continue
            seen.add(key)
            ordered.append(key)
    return ordered


def _column_label(key: str) -> str:
    spaced = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", key)
    spaced = spaced.replace("_", " ")
    return spaced[:1].upper() + spaced[1:]


def _serialise_cell(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)
